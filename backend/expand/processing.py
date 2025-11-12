import uuid

from backend.llm.openai import OpenAI, Prompt
from collections import defaultdict
from backend.config import config
import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import Font


def processing_v1_result(OpenAI: OpenAI, process_list: list, version = "v1") -> list:
    system_message = Prompt.get_system_message(version)
    # TODO: 实现对data_list的解析和返回
    grouped_data = defaultdict(list)
    for item in process_list:
        grouped_data[item['数据资源英文名称']].append(item)
    new_data_list = list(grouped_data.values())
    result_list = []
    for group in new_data_list:
        # 提取所有 *数据项英文名称 的值
        table = group[0]['数据资源英文名称']
        if table is None:
            continue
        data_item_names = [item['数据项英文名称'] for item in group]
        user_message = "\n".join(data_item_names)
        # TODO: 解析data
        # TODO: 调用OpenAI接口获取结果
        content = OpenAI.generate_with_system_message(system_message=system_message, user_message=user_message)
        json_content = content.content.strip().strip("```json").strip()
        json_content = json.loads(json_content)
        json_content['table'] = table
        result_list.append(json_content)
    return result_list  # TODO: 返回结果

def processing_v2_result(OpenAI: OpenAI, process_list: list, version = "v2") -> list:
    system_message = Prompt.get_system_message(version)
    # TODO: 实现对data_list的解析和返回
    grouped_data = defaultdict(list)
    for item in process_list:
        grouped_data[item['数据资源英文名称']].append(item)
    new_data_list = list(grouped_data.values())
    result_list = []
    for group in new_data_list:
        # 提取所有 *数据项英文名称 的值
        table = group[0]['数据资源英文名称']
        if table is None:
            continue
        data_item_names = [item['数据项英文名称'] for item in group]
        example_item_names = [str(item['示例数据']) for item in group]
        data_message = "\n".join(data_item_names)
        example_message = "\n".join(example_item_names)
        user_message = data_message +"\n 示例数据是:\n" + example_message
        # TODO: 解析data
        # TODO: 调用OpenAI接口获取结果
        content = OpenAI.generate_with_system_message(system_message=system_message, user_message=user_message)
        json_content = content.content.strip().strip("```json").strip()
        json_content = json.loads(json_content)
        json_content['table'] = table
        result_list.append(json_content)
    return result_list  # TODO: 返回结果


def generate_final_result(content: list, process_list: list) -> str:
    # TODO: 解析content，生成最终结果
    for data in process_list:
        for item in content:
            if item['table'] == data['数据资源英文名称']:
                data['*数据项中文名称\n（最多100个字）'] = item[data['数据项英文名称']]
    headers = list(process_list[0].keys())
    json_data = pd.DataFrame(process_list)
    json_data.columns = headers
    file_path = config.get('ai_excel_file') + f'{str(uuid.uuid4())}.xlsx'
    json_data.to_excel(file_path, index=False)
    return file_path

def generate_check_final_result(content: list, process_list: list) -> str:
    # TODO: 解析content，生成最终结果
    abnormal_list = []
    for process_index, data in enumerate(process_list):
        for item in content:
            if item['table'] == data['数据资源英文名称']:
                if data['*数据项中文名称\n（最多100个字）'] != item[data['数据项英文名称']]:
                    data['*数据项中文名称\n（最多100个字）'] = item[data['数据项英文名称']]
                    abnormal_list.append(process_index)
    headers = list(process_list[0].keys())
    json_data = pd.DataFrame(process_list)
    json_data.columns = headers
    file_path = config.get('ai_excel_file') + f'{str(uuid.uuid4())}.xlsx'
    json_data.to_excel(file_path, index=False)

    # 加载 Excel 文件
    wb = load_workbook(file_path)
    ws = wb.active
    # 设置特定单元格字体为红色，例如 A1 单元格
    for index in abnormal_list:
        ws['I' + str(index + 2)].font = Font(color="FF0000")
        ws['J' + str(index + 2)].font = Font(color="FF0000")
    wb.save(file_path)
    return file_path